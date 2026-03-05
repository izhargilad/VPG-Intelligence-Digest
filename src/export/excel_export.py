"""Excel export for VPG Intelligence Digest.

Generates .xlsx workbooks with multiple sheets:
- Signals: All scored signals with scores, BU, type
- Trends: Current trend data
- Keywords: Keyword performance
- Summary: Executive-level stats

Uses openpyxl for .xlsx generation.
Falls back gracefully if openpyxl is not installed.
"""

import logging
from datetime import datetime
from io import BytesIO
from pathlib import Path

from src.db import get_connection, get_all_industries, get_all_keywords

logger = logging.getLogger(__name__)


def _get_openpyxl():
    """Import openpyxl or return None."""
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        logger.warning("openpyxl not installed — Excel export disabled. pip install openpyxl")
        return None


def _style_header(ws, openpyxl):
    """Apply VPG branding to the header row."""
    from openpyxl.styles import Font, PatternFill, Alignment

    navy_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    white_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = navy_fill
        cell.font = white_font
        cell.alignment = center


def _auto_width(ws):
    """Auto-size columns based on content."""
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value or "")
            max_length = max(max_length, len(val))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)


def export_signals_excel(start_date: str | None = None, end_date: str | None = None,
                         bu_id: str | None = None, signal_type: str | None = None,
                         industry_id: str | None = None, min_score: float = 0,
                         output_path: Path | None = None) -> BytesIO | Path:
    """Export scored signals to an Excel workbook with optional filters.

    Args:
        start_date: Optional YYYY-MM-DD start filter.
        end_date: Optional YYYY-MM-DD end filter.
        bu_id: Filter by business unit.
        signal_type: Filter by signal type.
        industry_id: Filter by industry.
        min_score: Minimum composite score.
        output_path: If provided, saves to file and returns Path. Otherwise returns BytesIO.

    Returns:
        BytesIO buffer or Path to saved file.
    """
    openpyxl = _get_openpyxl()
    if openpyxl is None:
        raise ImportError("openpyxl is required for Excel export")

    from openpyxl.styles import Font, PatternFill, Alignment, numbers

    wb = openpyxl.Workbook()

    conn = get_connection()
    try:
        # ── Sheet 1: Signals ──
        ws_signals = wb.active
        ws_signals.title = "Signals"
        headers = ["ID", "Title", "Signal Type", "Headline", "Composite Score",
                    "Revenue Impact", "Time Sensitivity", "Strategic Alignment",
                    "Competitive Pressure", "Validation", "BUs", "Source", "URL",
                    "Published", "Quick Win", "Owner", "Est. Impact", "Source Links"]
        ws_signals.append(headers)

        query = """
            SELECT s.id, s.title, sa.signal_type, sa.headline, sa.score_composite,
                   sa.score_revenue_impact, sa.score_time_sensitivity,
                   sa.score_strategic_alignment, sa.score_competitive_pressure,
                   sa.validation_level, s.source_name, s.url, s.published_at,
                   sa.quick_win, sa.suggested_owner, sa.estimated_impact
            FROM signals s
            JOIN signal_analysis sa ON s.id = sa.signal_id
            WHERE s.status IN ('scored', 'published')
              AND COALESCE(s.dismissed, 0) = 0
        """
        params = []
        if start_date:
            query += " AND s.collected_at >= ?"
            params.append(start_date)
        if end_date:
            query += " AND s.collected_at <= ?"
            params.append(end_date + " 23:59:59")
        if signal_type:
            query += " AND sa.signal_type = ?"
            params.append(signal_type)
        if min_score > 0:
            query += " AND sa.score_composite >= ?"
            params.append(min_score)
        if bu_id:
            query += " AND s.id IN (SELECT signal_id FROM signal_bus WHERE bu_id = ?)"
            params.append(bu_id)
        if industry_id:
            query += " AND s.id IN (SELECT signal_id FROM signal_industries WHERE industry_id = ?)"
            params.append(industry_id)
        query += " ORDER BY sa.score_composite DESC"

        rows = conn.execute(query, params).fetchall()

        for row in rows:
            sig_id = row[0]
            bus = conn.execute(
                "SELECT bu_id FROM signal_bus WHERE signal_id = ?", (sig_id,)
            ).fetchall()
            bu_str = ", ".join(b[0] for b in bus)

            # Collect source links (primary URL + validation sources)
            source_links = [row[11]] if row[11] else []
            validations = conn.execute(
                "SELECT corroborating_url FROM signal_validations WHERE signal_id = ?",
                (sig_id,)
            ).fetchall()
            source_links.extend(v[0] for v in validations if v[0])
            links_str = " | ".join(source_links)

            ws_signals.append([
                sig_id, row[1], row[2], row[3],
                round(row[4] or 0, 1), round(row[5] or 0, 1), round(row[6] or 0, 1),
                round(row[7] or 0, 1), round(row[8] or 0, 1),
                row[9], bu_str, row[10], row[11], row[12],
                row[13], row[14], row[15], links_str,
            ])

        _style_header(ws_signals, openpyxl)
        _auto_width(ws_signals)

        # Color-code composite scores
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for r in range(2, ws_signals.max_row + 1):
            score_cell = ws_signals.cell(row=r, column=5)
            val = score_cell.value or 0
            if val >= 8.0:
                score_cell.fill = green_fill
            elif val >= 6.0:
                score_cell.fill = yellow_fill
            elif val >= 4.0:
                score_cell.fill = red_fill

        # ── Sheet 2: Trends ──
        ws_trends = wb.create_sheet("Trends")
        ws_trends.append(["Trend", "Type", "Momentum", "Signals", "WoW %",
                          "Avg Score", "Max Score", "First Seen", "Last Seen"])

        trend_rows = conn.execute("""
            SELECT label, trend_type, momentum, occurrence_count,
                   week_over_week_change, avg_score, max_score, first_seen, last_seen
            FROM trends ORDER BY occurrence_count DESC
        """).fetchall()

        for tr in trend_rows:
            ws_trends.append(list(tr))

        _style_header(ws_trends, openpyxl)
        _auto_width(ws_trends)

        # ── Sheet 3: Keywords ──
        ws_kw = wb.create_sheet("Keywords")
        ws_kw.append(["Keyword", "Industry", "Source", "Hit Count", "Active", "Last Hit"])

        keywords = get_all_keywords(conn, active_only=False)
        industries = {ind["id"]: ind["name"] for ind in get_all_industries(conn)}

        for kw in keywords:
            ws_kw.append([
                kw["keyword"],
                industries.get(kw.get("industry_id", ""), kw.get("industry_id", "")),
                kw.get("source", ""),
                kw.get("hit_count", 0),
                "Yes" if kw.get("active") else "No",
                kw.get("last_hit_at", ""),
            ])

        _style_header(ws_kw, openpyxl)
        _auto_width(ws_kw)

        # ── Sheet 4: Summary ──
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["VPG Intelligence Digest — Export Summary"])
        ws_summary.append([])
        ws_summary.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")])
        if start_date:
            ws_summary.append(["Start Date", start_date])
        if end_date:
            ws_summary.append(["End Date", end_date])
        if bu_id:
            ws_summary.append(["BU Filter", bu_id])
        if signal_type:
            ws_summary.append(["Signal Type Filter", signal_type])
        if industry_id:
            ws_summary.append(["Industry Filter", industry_id])
        if min_score > 0:
            ws_summary.append(["Min Score Filter", str(min_score)])
        ws_summary.append([])

        total_signals = conn.execute("SELECT COUNT(*) FROM signals").fetchone()[0]
        scored = conn.execute("SELECT COUNT(*) FROM signals WHERE status='scored'").fetchone()[0]
        pipeline_runs = conn.execute("SELECT COUNT(*) FROM pipeline_runs").fetchone()[0]

        ws_summary.append(["Metric", "Value"])
        ws_summary.append(["Total Signals", total_signals])
        ws_summary.append(["Scored Signals", scored])
        ws_summary.append(["Signals in Export", len(rows)])
        ws_summary.append(["Trends Tracked", len(trend_rows)])
        ws_summary.append(["Keywords Active", sum(1 for k in keywords if k.get("active"))])
        ws_summary.append(["Pipeline Runs", pipeline_runs])

        # Style the summary title
        title_font = Font(name="Arial", bold=True, size=14, color="1B2A4A")
        ws_summary["A1"].font = title_font

        _auto_width(ws_summary)

    finally:
        conn.close()

    # Output
    if output_path:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        logger.info("Excel export saved to %s", output_path)
        return output_path
    else:
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
